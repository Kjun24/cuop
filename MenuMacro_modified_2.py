import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from itertools import count

# -------------------------------------- #
# MOKI manager URL
url = "https://www.kioskmanager.co.kr/admin/ver2/login.php"
biz = 0
pw = 0

# Login
input_condition = True
while input_condition:
    print("사업자번호를 입력해주세요 : ")
    biz = input()
    if biz == 'test':
        biz = 4561237890    # 사업자번호
        pw = 1234567        # 비밀번호
        break
    print("입력하신 사업자번호는 [", biz, "] 입니다. ")
    print("맞으시면 Y, 틀리면 N을 입력해주세요.")
    correct_biznum = input()
    if correct_biznum == "Y" or correct_biznum == "y":
        print('비밀번호를 입력해주세요.')
        pw = input()
        input_condition = False
        print("비밀번호가 입력되었습니다. 메뉴등록을 시작합니다. 비밀번호가 틀렸을경우, 등록이 진행되지 않습니다.")
    elif correct_biznum == "N" or correct_biznum == "n":
        print('다시 시도합니다.')
    else:
        print('잘못 입력하셨습니다.')

NextNPrev_Menu_Gap = 17             # number of row (menu to menu)
First_Option_Row = 11               # 첫번째 옵션(옵션1)의 이름 등록(옵션명) 행 위치
Third_Option_Row = 17               # 세번째 옵션(옵션3)의 이름 등록(옵션명) 행 위치

# (여긴 손안대도 됩니당)세번째 옵션 행 위치 - 첫번째 옵션 행 위치 => 옵션이 몇개가 되든 양식에 규칙만 있다면 모두 추출 가능하도록.
Option_Row_Gap = Third_Option_Row - First_Option_Row

Max_Num_Option = 10
Max_Num_Contents = 10  


### selenium 세팅
# -------------------------------------- #
### 매크로, 페이지 로딩 최대 대기시간(1 = 1s)
macro_sleep = 0.8
loading_timeout = 1.5
def wait():
    driver.implicitly_wait(loading_timeout) # 웹 로딩 기다리는 최대시간 설정
    time.sleep(macro_sleep) # 행동간격 딜레이 설정


### 웹 불러온 뒤 로그인 세팅
def set_chrome_driver():
    chrome_options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    return driver
driver = set_chrome_driver()
driver.get(url)
driver.find_element("id","biz_imput").send_keys(biz)
driver.find_element("id","pw_input").send_keys(pw)
driver.find_element("id","login_btn").click()
wait()
# -------------------------------------- #

### opnpyxl / Open Excel File
xlsx_name = 'MENU_Input_Form_10.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
wb = openpyxl.load_workbook(filename=xlsx_name)
# -------------------------------------- #
CGName='D'; Menu='E'                            # CGName = 카테고리이름 // Menu = 메뉴명,가격
Op_Odd_S = 'D'; Op_Odd_P = 'G'                  # Op_Odd_S = 홀수옵션의 옵션명,옵션내용 // Op_Odd_P = 홀수옵션의 가격
Op_Even_S = 'J'; Op_Even_P = 'N'                # Op_Even_S = 짝수옵션의 옵션명,옵션내용 // Op_Even_P = 짝수옵션의 가격


Option_Column = ['E','F','G','H','I','E','F','G','H','I']           # [1,2,3,4,5]
Option_row = [11, 14]                           # it means 1st menu, 1st option, 1st content ... if you want to move to next option, add +6
Menu_to_Menu = 85                               # if you want to move to next menu, add +85
Menu_Info_Position = ['F5', 'F6', 'I5']         # [Menu Name, Menu Price, ETC]


### wb['엑셀시트이름']
sheet_list = ["카테고리1","카테고리2","카테고리3","카테고리4","카테고리5","카테고리6"]

def check_category(ws):
    if ws['E3'].value==None:  # 카테고리 이름 확인, 여기서 3은 양식에 있는 카테고리의 행 위치
        print("Empty Category.")
        return False
    else:
        driver.find_element("id","nav_plus_btn").click()
        wait()
        driver.find_element("name","ca_name").send_keys(ws['E3'].value)
        driver.find_element("id","add_dialog_confirm").click()
        wait()
        print("Category : "+ws['E3'].value)
        return True

def category_num(num):
    ### 카테고리 등록 주소로 이동
    driver.get("https://www.kioskmanager.co.kr/admin/ver2/category_in.php")
    wait()
    ws = wb[sheet_list[num-1]]

    # qid가 존재하지 않는지 확인 하고, 리스트에 없으면 해당 아이디에 등록하는 용도.
    div_list = []
    for i in range(6):
        try:
            qid = driver.find_element('xpath',('//*[@id="sortable"]/div['+str(i+1)+']')).get_attribute('data-id')
            if qid != None:
                div_list.append(qid)
        except NoSuchElementException:
            break
    # 카테고리 유무 확인
    if check_category(ws) == False:
        return
    
    # 행 값에 더하기 연산 변수 초기화.
    Row_Value_Plus = 0

    # 메뉴등록 시작 카테고리 N(작성기준 최대 6개의 카테고리)의 셀을 순차적으로 확인. 메뉴이름이 안보일때 까지.
    for i in range(6):
        qid = driver.find_element('xpath',('//*[@id="sortable"]/div['+str(i+1)+']')).get_attribute('data-id')
        if qid != None:
            if qid in div_list:
                print()
            else:
                print(qid+"가 없습니당.")
                div_list.append(qid)
                # ws[Menu,5] : 메뉴이름 확인, 여기서 5는 양식에 있는 메뉴이름의 행 위치.
                # 다음 메뉴를 확인할땐 다음 행 위치와 현재 행 위치의 차이를 합하여 값을 찾는다.(=Row_Value_Plus)
                while(1):
                    if ws['F'+str(5+Row_Value_Plus)].value==None:
                        print("메뉴의 내용이 없습니다. 다음 카테고리를 확인합니다.")
                        break
                    else:
                        # qid 및 자바스크립트를 이용하여 메뉴등록창 이동.
                        driver.execute_script(("location.href=('/admin/ver2/item_in.php?ca_id=" + str(qid) + "');"))
                        wait()
                        # 메뉴명 등록
                        driver.find_element('xpath','//*[@id="mid_right"]/div[1]/input').send_keys(ws['F'+str(5+Row_Value_Plus)].value)
                        # 중복확인 클릭 일단 중복 여부 무시하고 매크로 돌립니다.
                        driver.find_element('xpath','//*[@id="mid_right"]/div[1]/img').click()
                        wait()
                        print("메뉴명 : "+ws['F'+str(5+Row_Value_Plus)].value)
                    # ws[Menu,5] : 메뉴의 가격은 같은 열을 사용하기 때문에 메뉴와 변수동일, 여기서 7은 양식에 있는 가격의 행 위치.
                    if ws['F'+str(6+Row_Value_Plus)].value==None:
                        print("메뉴등록 실패 : 메뉴에 가격이 없어 현재 카테고리의 메뉴등록을 취소하고 다음 카테고리를 검색합니다.")
                        break
                    else:           # 가격 등록
                        driver.find_element('xpath','//*[@id="mid_right"]/div[2]/input').send_keys(ws['F'+str(6+Row_Value_Plus)].value)
                        print("Price : "+ str(ws['F'+str(6+Row_Value_Plus)].value))



                    Option_Number = 1
                    Option_row_count = 0
                    for options_index in range(10):
                        option_name_position = 9+Row_Value_Plus+(options_index*8)
                        if ws['E'+str(option_name_position)].value==None:
                            print("옵션명이 없습니다.")
                            break
                        else:# Adding Option Area 
                            driver.find_element('xpath','//*[@id="spl_subject'+str(int(Option_Number)-1)+'"]').send_keys(ws['E'+str(option_name_position)].value)
                            count = 0       #Number of Option Infos
                            while True:
                                option_info_position = Option_Column[count] + str(11+((count//5)*3)+Row_Value_Plus+options_index*8)
                                option_price_position = Option_Column[count] + str(12+((count//5)*3)+Row_Value_Plus+options_index*8)
                                if ws[option_info_position].value==None:
                                    break
                                else: # Put Option info (if there is no element, while roop rwill be stop)
                                    driver.find_element('xpath','//*[@id="spl_id'+str(int(Option_Number)-1)+'_'+str(count)+'"]').send_keys(ws[option_info_position].value)
                                if ws[option_price_position].value==None:
                                    driver.find_element('xpath','//*[@id="spl_price'+str(int(Option_Number)-1)+'_'+str(count)+'"]').send_keys('0')
                                else: # Put Option Prices (if there is no element, while roop will be stop)
                                    driver.find_element('xpath','//*[@id="spl_price'+str(int(Option_Number)-1)+'_'+str(count)+'"]').send_keys(str(ws[option_price_position].value))
                                    # Adding Option Button(+) script call
                                count = count+1
                                if count >= Max_Num_Contents:
                                    break
                            if options_index != 9:
                                driver.find_element('xpath','//*[@id="option_plus_btn_wrap"]/img').click()
                                wait()                
                        Option_Number += 1
                        if Option_Number > Max_Num_Option:
                            print("Enroll Menu Succeed")
                            break
                    Row_Value_Plus += Menu_to_Menu
                    driver.execute_script("item_add()")
                    wait()
                break
    return qid

# 매크로 시작
# 정렬을 위한 qid 보관 리스트
qid_list = []
for i in range(6):
    print("Category[" + str(i+1) + "] -> checking...")
    rtn = category_num(i+1)
    if rtn!=None:
        qid_list.append(rtn)

from category_sorting import category_sorting
category_sorting(biz,pw)

print("매크로를 종료합니다.")